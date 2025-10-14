import csv
import os
from io import TextIOWrapper, BytesIO, StringIO

from openpyxl import load_workbook, Workbook

CRM_MULTIFIELD = {'PHONE', 'EMAIL', 'WEB', 'IM', 'LINK'}


def iter_contacts(file):
    """
    Универсальный итератор контактов из файла.
    К файлу выдвигаются следующие требования:
    - указаны заголовки полей
    - наличие заголовка NAME
    - записи без значения в поле NAME игнорируются
    - значения для полей с множеством значений разделены запятой
    """
    ext = os.path.splitext(file.name.lower())[-1]
    # Допустимые расширения файла и их парсеры
    parsers = {
        ".csv": _parse_csv,
        ".xlsx": _parse_xlsx,
    }

    parser = parsers.get(ext)
    if not parser:
        raise ValueError(f"Неподдерживаемый формат файла: {ext}")

    for row in parser(file):
        # Пропуск строк без значения в NAME
        if not row.get("NAME"):
            continue
        # Преобразование multifield полей
        _parse_multifield(row)
        yield row


def _parse_multifield(row):
    """
    Преобразует поля контакта с множеством значений из строки в список словарей.
    Предпологается, что значения в строке разделены запятой.
    Также возможно, что строка содержит тип и значение, разделённые символом ':'.
    Например, "<значение 1>,<значение 2>,<тип значения 3>:<значение 3>".
    """
    for field in CRM_MULTIFIELD:
        if field in row:
            parsed_multifield = []
            if str(row[field]):
                for field_values in str(row[field]).split(','):
                    split_values = field_values.split(':', maxsplit=1)
                    if len(split_values) == 2:
                        value_type, value = split_values
                        parsed_multifield.append({'VALUE_TYPE': value_type, 'VALUE': value})
                    elif len(split_values) == 1:
                        value = split_values[0]
                        parsed_multifield.append({'VALUE': value})
            row[field] = parsed_multifield


def _parse_csv(file, delimiter=";"):
    file.seek(0)
    wrapper = TextIOWrapper(file.file, encoding="utf-8")
    reader = csv.reader(wrapper, delimiter=delimiter)

    try:
        headers = next(reader)
    except StopIteration:
        return  # Пустой файл

    headers = [h.strip() for h in headers if h is not None]
    if "NAME" not in headers:
        return  # Если нет столбца NAME - не парсим

    for row in reader:
        row_data = {}
        for i, header in enumerate(headers):
            value = str(row[i]).strip() if i < len(row) else ""
            row_data[header] = value or ""
        yield row_data


def _parse_xlsx(file):
    file.seek(0)
    data = BytesIO(file.read())
    wb = load_workbook(data, read_only=True, data_only=True)
    sheet = wb.active

    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h else "" for h in next(rows)]
    except StopIteration:
        return  # Пустой файл

    if "NAME" not in headers:
        return  # Если нет столбца NAME - не парсим

    for values in rows:
        row_data = {}
        for i, header in enumerate(headers):
            value = values[i] if i < len(values) else ""
            if isinstance(value, str):
                value = value.strip()
            row_data[header] = value or ""
        yield row_data


def write_contacts(contacts, headers=None, ext='.csv'):
    """
    Универсальная функция экспорта контактов в файл, возвращает объект BytesIO.
    Первая строка файла - заголовки.
    """
    writers = {
        ".csv": _to_csv,
        ".xlsx": _to_xlsx,
    }

    writer = writers.get(ext.lower())
    if not writer:
        raise ValueError(f"Неподдерживаемый формат файла: {ext}")

    contacts = list(contacts)
    if not contacts:
        raise ValueError("Нет данных для записи")

    if headers is None:
        headers = list({key for c in contacts for key in c.keys()})

    return writer(contacts, headers)


def _to_multifield(value):
    """Преобразует значения multifield в строку значений через запятую."""
    if not value:
        # Если нет значения
        return ""
    if isinstance(value, str):
        # Если значение уже строка
        return value
    if isinstance(value, list):
        parts = []
        for item in value:
            if not isinstance(item, dict):
                continue
            val = str(item.get("VALUE", "")).strip()
            """ Преобразование с типом значения
            val_type = item.get("VALUE_TYPE")
            if val_type:
                parts.append(f"{val_type}:{val}")
            else:
                parts.append(val)
            """
            parts.append(val)
        return ",".join(parts)
    return str(value)


def _to_csv(contacts, headers, delimiter=";"):
    buffer = StringIO()
    writer = csv.writer(buffer, delimiter=delimiter)

    writer.writerow(headers)

    for contact in contacts:
        row = []
        for h in headers:
            value = contact.get(h, "")
            if h in CRM_MULTIFIELD:
                value = _to_multifield(value)
            row.append(value)
        writer.writerow(row)

    byte_buffer = BytesIO(buffer.getvalue().encode("utf-8"))
    byte_buffer.seek(0)
    return byte_buffer


def _to_xlsx(contacts, headers):
    wb = Workbook()
    ws = wb.active

    ws.append(headers)

    for contact in contacts:
        row = []
        for h in headers:
            value = contact.get(h, "")
            if h in CRM_MULTIFIELD:
                value = _to_multifield(value)
            row.append(value)
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
